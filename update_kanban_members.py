import openpyxl

file_path = 'kanban/Kerne_Weekend_Kanban_Feb21_22.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Find the column index for 'Members' and 'Task ID'
header_row = 1
members_col = None
task_id_col = None

for col in range(1, ws.max_column + 1):
    val = ws.cell(row=header_row, column=col).value
    if val == 'Members':
        members_col = col
    elif val == 'Task ID':
        task_id_col = col

assignments = {
    1: 'Scofield',
    2: 'Scofield',
    3: 'Bagwell',
    4: 'Mahone',
    5: 'Scofield',
    6: 'Scofield',
    7: 'Bagwell',
    8: 'Mahone',
    9: 'Scofield',
    10: 'Scofield',
    11: 'Bagwell',
    12: 'Scofield',
    13: 'Abruzzi',
    14: 'Mahone',
    15: 'Abruzzi',
    16: 'Scofield',
    17: 'Abruzzi',
    18: 'Scofield',
    19: 'Scofield',
    20: 'Mahone',
    21: 'Abruzzi',
    22: 'Mahone',
    23: 'Bagwell',
    24: 'Abruzzi',
    25: 'Mahone',
    26: 'Abruzzi',
    27: 'Mahone',
    28: 'Mahone',
    29: 'Scofield',
    30: 'Bagwell',
    31: 'Bagwell',
    32: 'Mahone',
    33: 'Scofield'
}

for row in range(2, ws.max_row + 1):
    task_id_cell = ws.cell(row=row, column=task_id_col)
    if task_id_cell.value in assignments:
        ws.cell(row=row, column=members_col).value = assignments[task_id_cell.value]

wb.save(file_path)
print("Updated successfully.")