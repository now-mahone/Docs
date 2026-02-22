import subprocess
import sys

def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
    import openpyxl
except ImportError:
    install('openpyxl')
    import openpyxl

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# Create workbook and select active sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Kerne Kanban"

# Define headers
headers = ["Task ID", "Members", "Task", "Status", "Notes", "Why we do it", "How", "What we gain from it", "Worst case scenario"]
ws.append(headers)

# Define tasks data
raw_tasks = [
    [
        "Modularize YRE Adapters",
        "Isolates integrations so external failures don't cascade. Allows rapid addition of new yield sources.",
        "Define a universal `IYieldAdapter` interface and rewrite existing strategy adapters to conform to it.",
        "Highly scalable architecture; adding new strategies takes days instead of weeks.",
        "Interface is too rigid for complex sources. Mitigate with flexible data payloads."
    ],
    [
        "Dynamic Hedge Rebalancing Engine V2",
        "Maximizes capital efficiency and minimizes delta exposure by reacting instantly to volatility.",
        "Implement WebSocket listeners for vault events and integrate with CCXT exchange managers.",
        "Near-perfect delta neutrality with reduced slippage and gas costs.",
        "Logic bug causes over-hedging. Mitigate with hard-coded circuit breakers."
    ],
    [
        "Scale Autonomous Institutional Outreach",
        "Builds a massive TVL pipeline automatically, freeing up time to focus on closing deals.",
        "Refine LLM prompts for personalized messaging and integrate LinkedIn/governance forum data.",
        "Consistent stream of warm institutional leads and early strategic partnerships.",
        "Bot sends hallucinated messages. Mitigate with a human-in-the-loop approval step."
    ],
    [
        "Finalize Privacy Policy & ToS",
        "Protects the team from liability and clearly defines Kerne as software, not a financial institution.",
        "Review existing drafts for cross-chain/DAO edge cases and integrate mandatory acceptance flows.",
        "Robust legal shield that deters lawsuits and provides clarity to institutional users.",
        "Loophole exposes us to liability. Mitigate by adhering to established DeFi legal templates."
    ],
    [
        "Cross-Chain Messaging Abstraction",
        "Removes single points of failure. Allows seamless switching or multi-routing of messages.",
        "Design an `IMessageRelay` interface and implement adapters for major messaging protocols.",
        "Unparalleled resilience against bridge hacks, ensuring protocol solvency across chains.",
        "Vulnerability in the abstraction layer. Mitigate through rigorous formal verification."
    ],
    [
        "Automated Risk Scoring Oracle",
        "Quantifies risk in real-time to safely route capital and rapidly exit deteriorating strategies.",
        "Develop a service aggregating DeFiLlama/security data to calculate scores and update on-chain caps.",
        "Institutional-grade risk management preventing catastrophic losses from external exploits.",
        "Oracle misses a novel exploit. Mitigate by maintaining conservative allocation caps."
    ],
    [
        "Pre-Deposit 'Points' Campaign",
        "Bootstraps initial TVL and community momentum, allowing us to gauge actual market demand.",
        "Design point emission logic, deploy a secure escrow contract, and build a frontend leaderboard.",
        "Tens of millions in committed TVL before launch, ensuring deep peg stability and social proof.",
        "Escrow contract vulnerability. Mitigate with extremely simple logic and multiple audits."
    ],
    [
        "Multi-Sig Operational Security",
        "Eliminates single points of failure in key management, protecting against compromised keys.",
        "Set up Safe contracts with a 3-of-4 threshold using hardware wallets and document strict procedures.",
        "Security posture meeting institutional requirements, eliminating key-person risk.",
        "Loss of multiple hardware wallets. Mitigate with geographically distributed encrypted backups."
    ],
    [
        "Real-Time On-Chain Analytics",
        "Radical transparency builds unbreakable trust and proves solvency to data-driven investors.",
        "Write SQL queries on Dune Analytics and build a dedicated 'Transparency' page on the frontend.",
        "Powerful marketing asset proving kUSD is fully backed and generating sustainable yield.",
        "Data indexing error causes panic. Mitigate by cross-referencing data and adding disclaimers."
    ],
    [
        "Optimize ZIN Solver",
        "Maximizes spread capture and revenue by accessing the deepest liquidity and highest intent volume.",
        "Refactor solver for modular venue adapters and update settlement logic for specific platform needs.",
        "Significantly increased daily protocol revenue, establishing Kerne as a dominant solver.",
        "Settlement logic vulnerability leads to lost gas. Mitigate by testing on mainnet forks."
    ],
    [
        "Co-Marketing Playbooks",
        "Ensures maximum TVL and brand awareness extraction from every partnership.",
        "Draft a playbook detailing timelines for announcements, Spaces, blogs, and shared campaigns.",
        "Massive multiplier effect on user acquisition by leveraging partners' existing audiences.",
        "Partner executes poorly, causing brand confusion. Mitigate by requiring mutual approval."
    ],
    [
        "ZK-Proof Solvency Attestation",
        "Bridges the trust gap between on-chain transparency and off-chain hedging without revealing secrets.",
        "Integrate a ZK-coprocessor or oracle network to verify exchange balances via read-only APIs.",
        "Neutralizes 'FTX risk' narrative, unlocking capital from security-conscious institutions.",
        "Proof generation fails during volatility. Mitigate with redundant attestation methods."
    ],
    [
        "Ambassador Program Framework",
        "Organic growth is more sustainable than paid ads. Allows rapid expansion into new markets.",
        "Create a tiered reward system (KERNE tokens), set up Discord channels, and track monthly KPIs.",
        "Decentralized marketing army generating continuous social proof and user onboarding.",
        "Ambassadors use spammy tactics. Mitigate with strict brand guidelines and zero-tolerance policies."
    ],
    [
        "Cross-Chain Liquidity Monitoring",
        "Deep liquidity is essential for peg stability. Proactive monitoring prevents depeg events.",
        "Write a script querying DEX subgraphs for balances and integrate with an AlertManager.",
        "Rock-solid peg across all chains, preventing arbitrageurs from exploiting thin order books.",
        "Monitoring fails silently. Mitigate by running redundant instances and daily health checks."
    ],
    [
        "Institutional Client Documentation",
        "Institutions require detailed risk and legal info. Ready documentation speeds up large allocations.",
        "Write detailed PDFs covering architecture, hedging, risk parameters, and custody integrations.",
        "Drastically reduced sales cycle for institutional clients, signaling competence and reliability.",
        "Documentation becomes outdated. Mitigate by integrating updates into the deployment pipeline."
    ],
    [
        "MEV Protection Layer",
        "Protects execution, preserving capital efficiency and retaining maximum yield for users.",
        "Route transactions through private RPCs (Flashbots) and implement slippage/deadline checks.",
        "Saves hundreds of thousands in lost value, translating to higher APYs for kUSD holders.",
        "Private RPC fails during liquidation. Mitigate with fallback logic to public mempools."
    ],
    [
        "Yield Mechanics Educational Content",
        "Demystifies complex strategies, building trust and overcoming skepticism of yield-bearing stables.",
        "Break down complex strategies into understandable concepts and publish across multiple platforms.",
        "Establishes Kerne as a thought leader, continuously converting skeptical users into depositors.",
        "Oversimplification leads to inaccuracy. Mitigate by having engineering review all content."
    ],
    [
        "Decentralized Keeper Network",
        "Removes centralized points of failure, ensuring protocol autonomy if internal infrastructure fails.",
        "Write resolver contracts defining execution conditions and configure tasks on the decentralized network.",
        "True protocol autonomy and censorship resistance, strengthening regulatory positioning.",
        "Network outage prevents critical liquidations. Mitigate by keeping internal bots as a fallback."
    ],
    [
        "Tiered Liquidity Mining Structure",
        "Flat schedules attract mercenary farmers. Tiers acquire sticky, long-term TVL that stabilizes Kerne.",
        "Model emission curves and multiplier tiers, then implement them into staking smart contracts.",
        "Highly loyal depositor base and reduced circulating supply, creating upward price pressure.",
        "Lockups are too restrictive. Mitigate by offering a liquid option with base emissions."
    ],
    [
        "Incident Response Playbook",
        "Prevents panic and indecision. Ensures rapid execution of emergency pauses and communication.",
        "Draft emergency contacts, communication templates, and technical steps for pausing contracts.",
        "Drastically reduced response times, potentially saving millions and maintaining community trust.",
        "Playbook is too rigid. Mitigate by designing it as a flexible framework, empowering judgment calls."
    ],
    [
        "'Kerne Alpha' Twitter Spaces",
        "Builds parasocial connections, shapes the narrative, and cross-pollinates audiences with partners.",
        "Schedule Spaces, secure guests from partner protocols, co-host, and repurpose audio into clips.",
        "Rapid Twitter growth, establishing Kerne as a DeFi alpha hub and converting listeners to users.",
        "Accidental regulatory misstep live. Mitigate with clear talking points and avoiding price talk."
    ],
    [
        "Optimize Docker & Serverless",
        "Ensures infrastructure handles massive traffic spikes without latency as TVL grows.",
        "Refactor Yield Server to use Edge Functions and streamline Dockerfiles for faster bot deployment.",
        "99.99% uptime and sub-second response times, ensuring flawless UX and hedging operation.",
        "Migration misconfiguration causes an outage. Mitigate with blue-green deployments and staging."
    ],
    [
        "RWA Collateral Target List",
        "Diversifying into RWAs provides a stable yield floor during crypto bear markets.",
        "Analyze top RWA protocols based on TVL, legal structure, and security, then initiate BD talks.",
        "Positions Kerne as a bridge between TradFi yields and DeFi, attracting risk-averse capital.",
        "Integrating a non-compliant RWA. Mitigate by strictly limiting caps and partnering with compliant issuers."
    ],
    [
        "Discord/Telegram Gamification",
        "Engaged communities defend against competitors and can be mobilized for marketing/governance.",
        "Configure tools to track activity, assign roles/XP, and tie them to tangible protocol benefits.",
        "Vibrant, self-sustaining community that educates new users and reduces the support burden.",
        "Incentivizes spam. Mitigate by rewarding quality over quantity and empowering human moderators."
    ],
    [
        "Weekly 'State of Kerne' Reports",
        "Radical transparency maintains trust during downturns, showing exactly what's under the hood.",
        "Automate data extraction from dashboards and format it into a digestible newsletter with commentary.",
        "Unparalleled reputation for honesty, becoming required reading for institutional investors.",
        "Report shows temporary underperformance. Mitigate by providing clear context and corrective actions."
    ],
    [
        "Beta Tester Feedback Loop",
        "Ensures we build a product people want to use, allowing rapid UI/UX pivots before mainnet.",
        "Set up private channels/forms for testers and review feedback weekly to prioritize improvements.",
        "Highly polished, intuitive product that reduces friction during the critical initial TVL ramp-up.",
        "Overwhelmed by conflicting feedback. Mitigate by ruthlessly prioritizing based on strategic goals."
    ],
    [
        "Decentralized Risk Committee",
        "Distributes power, satisfying regulatory requirements and empowering the community.",
        "Draft a governance mandate defining elections, compensation, and slashing mechanics for bad approvals.",
        "Scalable risk management leveraging community intelligence, allowing rapid strategy integration.",
        "Committee captured by malicious actors. Mitigate by requiring high stakes and DAO veto power."
    ],
    [
        "Refine Invariant Testing Strategy",
        "Mathematical certainty of core logic is non-negotiable for confidently deploying upgrades.",
        "Write specific, bounded invariant properties and configure continuous fuzzing campaigns.",
        "Highest possible smart contract security, reducing external audit costs and exploit probability.",
        "Fuzzing misses a complex vulnerability. Mitigate by combining with manual audits and bug bounties."
    ],
    [
        "ZIN Tokenomics Smart Contracts",
        "Tokenomics are the economic engine. Flawless implementation aligns incentives and drives growth.",
        "Write the ERC-20 contract, timelocked vesting wallets, and the revenue fee-splitter contract.",
        "Robust economic system creating persistent buy pressure and maximizing fully diluted valuation.",
        "Bug allows claiming unearned tokens. Mitigate by using OpenZeppelin templates and formal verification."
    ],
    [
        "Lending Protocol Integrations",
        "Utility drives demand. Lending integration allows users to leverage kUSD, creating massive organic demand.",
        "Draft governance proposals highlighting kUSD's liquidity/safety and engage delegates for votes.",
        "Unlocks billions in potential TVL through 'double dip' utility, making kUSD highly attractive.",
        "Proposal rejected due to politics. Mitigate by building delegate relationships beforehand."
    ],
    [
        "LST/LRT Provider Partnerships",
        "Kerne amplifies LST demand. Partnerships allow joint campaigns tapping into massive user bases.",
        "Initiate BD conversations proposing boosted emissions for their collateral in exchange for promotion.",
        "Rapid acquisition of high-quality TVL by aligning with the biggest players in the staking ecosystem.",
        "Partner protocol suffers an exploit. Mitigate by strictly capping exposure to any single LST/LRT."
    ],
    [
        "Emergency Pause Mechanisms",
        "Seconds matter during an exploit. Granular pauses halt malicious activity without full shutdowns.",
        "Integrate OpenZeppelin's `Pausable` extension and define roles for the security multisig.",
        "Ultimate safety net preventing total capital loss, mandatory for top-tier audits and institutions.",
        "Pause abused by compromised admin. Mitigate with geographically distributed multi-sig control."
    ],
    [
        "Internal Smart Contract Fuzzing",
        "Finds edge cases before hackers do, allowing quiet fixes before paying external auditors.",
        "Write extensive property-based tests defining core invariants and run them continuously.",
        "Highly secure codebase entering audits, speeding up review timelines and lowering exploit probability.",
        "Fuzzing consumes excessive time without finding bugs. Mitigate by focusing on critical functions."
    ]
]

# Add data
for i, task in enumerate(raw_tasks, 1):
    # Task ID, Members, Task, Status, Notes, Why we do it, How, What we gain from it, Worst case scenario
    row = [
        i,
        "", # Members
        task[0], # Task
        "Not started", # Status
        "", # Notes
        task[1], # Why
        task[2], # How
        task[3], # Gain
        task[4]  # Worst case
    ]
    ws.append(row)

# Formatting
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
wrap_alignment = Alignment(vertical="top", wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Apply header formatting
for col_num, cell in enumerate(ws[1], 1):
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = thin_border

# Apply data formatting
for row in ws.iter_rows(min_row=2, max_row=len(raw_tasks)+1):
    for cell in row:
        cell.alignment = wrap_alignment
        cell.border = thin_border
        
    # Center align Task ID
    row[0].alignment = Alignment(horizontal="center", vertical="top")

# Set column widths
column_widths = {
    "A": 10, # Task ID
    "B": 15, # Members
    "C": 35, # Task
    "D": 15, # Status
    "E": 25, # Notes
    "F": 40, # Why we do it
    "G": 40, # How
    "H": 40, # What we gain from it
    "I": 40  # Worst case scenario
}
for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Data Validation for Members
members_dv = DataValidation(type="list", formula1='"Scofield,Mahone,Bagwell,Abruzzi,Everyone"', allow_blank=True)
ws.add_data_validation(members_dv)
members_dv.add(f"B2:B{len(raw_tasks)+1}")

# Data Validation for Status
status_dv = DataValidation(type="list", formula1='"Not started,Started,Ongoing,Complete"', allow_blank=True)
ws.add_data_validation(status_dv)
status_dv.add(f"D2:D{len(raw_tasks)+1}")

# Conditional Formatting for Members Colors
# Abruzzi = Yellow, Scofield = Blue, Mahone = Green, Bagwell = Orange, Everyone = Grey
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
blue_fill = PatternFill(start_color="C9DAF8", end_color="C9DAF8", fill_type="solid")
green_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
orange_fill = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
grey_fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")

ws.conditional_formatting.add(f"B2:B{len(raw_tasks)+1}", CellIsRule(operator='equal', formula=['"Abruzzi"'], fill=yellow_fill))
ws.conditional_formatting.add(f"B2:B{len(raw_tasks)+1}", CellIsRule(operator='equal', formula=['"Scofield"'], fill=blue_fill))
ws.conditional_formatting.add(f"B2:B{len(raw_tasks)+1}", CellIsRule(operator='equal', formula=['"Mahone"'], fill=green_fill))
ws.conditional_formatting.add(f"B2:B{len(raw_tasks)+1}", CellIsRule(operator='equal', formula=['"Bagwell"'], fill=orange_fill))
ws.conditional_formatting.add(f"B2:B{len(raw_tasks)+1}", CellIsRule(operator='equal', formula=['"Everyone"'], fill=grey_fill))

# Conditional Formatting for Status Colors
status_not_started = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")
status_started = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
status_ongoing = PatternFill(start_color="C9DAF8", end_color="C9DAF8", fill_type="solid")
status_complete = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

ws.conditional_formatting.add(f"D2:D{len(raw_tasks)+1}", CellIsRule(operator='equal', formula=['"Not started"'], fill=status_not_started))
ws.conditional_formatting.add(f"D2:D{len(raw_tasks)+1}", CellIsRule(operator='equal', formula=['"Started"'], fill=status_started))
ws.conditional_formatting.add(f"D2:D{len(raw_tasks)+1}", CellIsRule(operator='equal', formula=['"Ongoing"'], fill=status_ongoing))
ws.conditional_formatting.add(f"D2:D{len(raw_tasks)+1}", CellIsRule(operator='equal', formula=['"Complete"'], fill=status_complete))

# Freeze the top row
ws.freeze_panes = "A2"

wb.save("Kerne_Kanban_Board.xlsx")
print("Successfully generated Kerne_Kanban_Board.xlsx")